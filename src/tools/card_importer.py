#- encoding: utf-8 -
#/*
#@File     :   card_importer.py
#@Desc     :   从骰子工厂格式 Excel 角色卡导入调查员角色
#@Note     :   仅支持当前已知的 .xlsx 布局；衍生值由系统自动重算
#             技能命名规则：四项冲突名映射到系统名，编号技能保留原名
#*/

from __future__ import annotations

import logging
import re
from dataclasses import asdict
from pathlib import Path
from typing import Any, Optional

from src.domain.character import (
    Character, Stats,
    create_investigator,
)

logger = logging.getLogger(__name__)


# ------- 角色卡仓库路径 -------

# 存放 .xlsx 角色卡文件的默认目录，自动创建
CARDS_DIR = Path(__file__).resolve().parent.parent.parent / "data" / "cards"


def ensure_cards_dir() -> Path:
    """确保 cards 目录存在并返回路径"""
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    return CARDS_DIR


# 旧版 data/card/ 目录（兼容降级搜索）
_LEGACY_CARD_DIR = CARDS_DIR.parent / "card"


def search_cards_dir(keyword: str) -> list[Path]:
    """在 cards 目录中搜索匹配关键字的 .xlsx 文件

    匹配规则：忽略大小写、忽略空格，文件名包含关键字即可。
    搜索顺序：data/cards/ 优先 → data/card/ 降级。

    Returns:
        匹配的文件路径列表（按文件名排序）
    """
    ensure_cards_dir()
    kw = keyword.lower().replace(" ", "")
    hits: list[Path] = []

    def _search(root: Path) -> list[Path]:
        if not root.is_dir():
            return []
        found: list[Path] = []
        for f in sorted(root.iterdir()):
            if f.suffix.lower() in (".xlsx", ".xls") and kw in f.stem.lower().replace(" ", ""):
                found.append(f)
        return found

    hits = _search(CARDS_DIR)
    if not hits and _LEGACY_CARD_DIR.is_dir():
        hits = _search(_LEGACY_CARD_DIR)
    return hits


def is_path_like(text: str) -> bool:
    """判断输入是否像文件路径（含分隔符或扩展名）"""
    return any(c in text for c in ("/", "\\", "."))


# ------- 技能名映射表 -------

# 四项冲突名按用户要求显式映射；编号技能（科学①、外语①、技艺①等）
# 和 Ω 后缀名保留原名，让运行时子串匹配兜底

SKILL_NAME_MAP: dict[str, str] = {
    "图书馆使用": "图书馆利用",
    "母语": "语言(母语)",
    "驯兽": "动物驯养",
}

_STRIP_SUFFIX_RE = re.compile(r"[Ω：\u3000\s]+$")


def _normalize(raw: Any) -> str:
    """标准化技能名：去前后空白 → 映射 → 去 Ω/全角冒号后缀"""
    if not raw:
        return ""
    name = str(raw).strip()
    mapped = SKILL_NAME_MAP.get(name)
    if mapped:
        return mapped
    return _STRIP_SUFFIX_RE.sub("", name)


def _to_int(v: Any) -> int:
    """安全转 int，非数字返回 0"""
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (ValueError, AttributeError, TypeError):
        return 0


# ------- 人物卡 Sheet 坐标常量 -------

# 以 openpyxl 1-indexed 行列表示
_ROW_BASIC = 3        # 姓名、STR/DEX/POW
_ROW_JOB = 5          # 职业、CON/APP/EDU
_ROW_SIZ_INT = 7      # SIZ/INT/Luck
_ROW_POOL = 10        # HP/SAN/MP
_ROW_DERIVED = 9      # DB/Build/MOV
_ROW_SKILL_START = 16 # 技能表起始行
_ROW_SKILL_END = 54   # 技能表结束行
_ROW_BG_START = 60    # 背景故事起始行
_ROW_BG_END = 77      # 背景故事结束行
_ROW_INV_START = 78   # 背包起始
_ROW_INV_END = 91     # 背包结束
_ROW_SPELL_HEADER = 112 # 法术区域
_COL_NAME = 5         # E 列
_COL_GENDER_AGE = 13  # M 列（性别/年龄存放位置）
# 属性坐标：每行三个属性，label/value/half 各间隔 6 列
_STAT_COLS = [
    ("STR", 21, 3),    # U3
    ("DEX", 27, 3),    # AA3
    ("POW", 33, 3),    # AG3
    ("CON", 21, 5),    # U5
    ("APP", 27, 5),    # AA5
    ("EDU", 33, 5),    # AG5
    ("SIZ", 21, 7),    # U7
    ("INT", 27, 7),    # AA7
]
_LUCK_COL = 33         # AG
_LUCK_ROW = 7

# 技能表：左半 (F/J/L/N/P) 和 右半 (AB/AF/AH/AJ/AL)
_SKILL_LEFT = dict(name=6, init=10, growth=12, occ=14, interest=16)
_SKILL_RIGHT = dict(name=28, init=32, growth=34, occ=36, interest=38)

# 背景：AA 列 (col=27) 存值  /  W 列 (col=23) 存标签
_BG_COL = 27
_BG_LABEL_COL = 23
_BG_ROWS: dict[str, int] = {
    "appearance_desc": 61,
    "belief": 63,
    "significant_person": 65,
    "significant_place": 67,
    "cherished_possession": 69,
    "trait": 71,
    "injury_scar": 73,
}
# 恐惧症和躁狂症（AA75，可能为空）
_PHOBIA_ROW = 75
# 完整背景故事（W77，小说式段落）
_FULL_BG_ROW = 77
_FULL_BG_COL = 23

# HP/SAN/MP 列坐标
_HP_CUR = 5   # E10
_HP_MAX = 7   # G10
_SAN_CUR = 14 # N10
_SAN_MAX = 16 # P10
_MP_CUR = 23  # W10
_MP_MAX = 25  # Y10
_MOV = 32     # AF10

# 法术：Y(col=25)=名称, AC(col=29)=代价, AH(col=34)=效果
_SPELL_COL_NAME = 25
_SPELL_COST = 29
_SPELL_EFFECT = 34
_SPELL_START = 114

# 背包：F(col=6)/N(col=14) 列
_INV_COL_LEFT = 6
_INV_COL_RIGHT = 14


# ------- 公共入口 -------

def import_from_xlsx(filepath: str | Path) -> Character:
    """解析 xlsx 角色卡，返回系统 Character 对象。

    解析流程：先读基本信息与属性 → 再读技能表并映射名 →
    读背景/背包/法术 → 调用 create_investigator 生成角色
    （衍生值自动重算，无视卡中数据），最后覆盖 Luck 保留原值。

    Args:
        filepath: .xlsx 文件路径

    Returns:
        Character: 含完整属性与技能数据的调查员角色
    """
    import openpyxl
    wb = openpyxl.load_workbook(str(filepath), data_only=True)
    ws = wb["人物卡"]

    # 先提取所有原始数据，再组装
    name, gender, age, birthplace, occ_name = _parse_basic_info(ws)
    stats, luck_value = _parse_stats(ws)
    skill_map = _parse_skills(ws)
    bg = _parse_background(ws)
    spells = _parse_spells(ws)
    inventory = _parse_inventory(ws)
    full_backstory = _parse_full_backstory(ws)
    phobias = _parse_phobias(ws)

    # 衍生值由系统自动计算，忽视卡中数据
    char = create_investigator(
        name=name,
        occupation=occ_name,
        stats=stats,
        occupation_skills=skill_map,
        gender=gender,
        age=age,
        birthplace=birthplace,
        spells=spells,
        **bg,
    )

    # Luck 是随机衍生的，覆盖为卡中原值保留原始骰运
    char.luck = luck_value
    char.inventory = inventory
    char.full_backstory = full_backstory
    char.phobias_manias = phobias

    # 同步 max_hit_points/max_magic_points（create_investigator 已算对）
    char.hit_points = char.max_hit_points
    char.magic_points = char.max_magic_points

    logger.info("角色卡导入完成: %s (%s) STR=%d DEX=%d ...",
                char.name, char.occupation,
                char.stats.strength, char.stats.dexterity)
    return char


# ------- 基本信息 -------

def _parse_basic_info(ws) -> tuple[str, str, int, str, str]:
    """提取姓名、性别、年龄、出生地、职业"""
    name = str(ws.cell(row=3, column=_COL_NAME).value or "").strip()

    # M6 是性别（本卡格式：M6="女"）
    gender_raw = ws.cell(row=6, column=_COL_NAME + 8).value  # M6
    gender = str(gender_raw or "").strip()

    # E6 是年龄（本卡格式：E6=24）
    age = _to_int(ws.cell(row=6, column=_COL_NAME).value)

    # E7 住地
    birthplace = str(ws.cell(row=7, column=_COL_NAME).value or "").strip()

    # E5 职业
    occ_name = str(ws.cell(row=5, column=_COL_NAME).value or "").strip()

    logger.debug("基本信息: name=%s gender=%s age=%d occ=%s",
                 name, gender, age, occ_name)
    return name, gender, age, birthplace, occ_name


# ------- 属性 -------

def _parse_stats(ws) -> tuple[Stats, int]:
    """提取八项属性和幸运值"""
    kwargs: dict[str, int] = {}
    luck = 0
    for label, col, row in _STAT_COLS:
        v = _to_int(ws.cell(row=row, column=col).value)
        key = {
            "STR": "strength", "CON": "constitution", "SIZ": "size",
            "DEX": "dexterity", "APP": "appearance", "INT": "intelligence",
            "POW": "power", "EDU": "education",
        }.get(label, "")
        if key:
            kwargs[key] = v
    luck = _to_int(ws.cell(row=_LUCK_ROW, column=_LUCK_COL).value)

    logger.debug("属性: %s Luck=%d", kwargs, luck)
    return Stats(**kwargs), luck


# ------- 技能表 -------

def _parse_skills(ws) -> dict[str, int]:
    """解析技能表，合并 init+growth+occ+interest 为最终值

    布局：左半 col F(名称)/J(初始)/L(成长)/N(职业)/P(兴趣)
          右半 col AB(名称)/AF(初始)/AH(成长)/AJ(职业)/AL(兴趣)
    """
    result: dict[str, int] = {}

    for r in range(_ROW_SKILL_START, _ROW_SKILL_END + 1):
        # 左半
        name_raw = ws.cell(row=r, column=_SKILL_LEFT["name"]).value
        if name_raw:
            name = _normalize(name_raw)
            if name:
                init = _to_int(ws.cell(row=r, column=_SKILL_LEFT["init"]).value)
                growth = _to_int(ws.cell(row=r, column=_SKILL_LEFT["growth"]).value)
                occ = _to_int(ws.cell(row=r, column=_SKILL_LEFT["occ"]).value)
                interest = _to_int(ws.cell(row=r, column=_SKILL_LEFT["interest"]).value)
                total = init + growth + occ + interest
                if total > 0 or not result:  # 总是保留第一个技能确保占位
                    result[name] = total

        # 右半
        name_raw = ws.cell(row=r, column=_SKILL_RIGHT["name"]).value
        if name_raw:
            name = _normalize(name_raw)
            if name:
                init = _to_int(ws.cell(row=r, column=_SKILL_RIGHT["init"]).value)
                growth = _to_int(ws.cell(row=r, column=_SKILL_RIGHT["growth"]).value)
                occ = _to_int(ws.cell(row=r, column=_SKILL_RIGHT["occ"]).value)
                interest = _to_int(ws.cell(row=r, column=_SKILL_RIGHT["interest"]).value)
                total = init + growth + occ + interest
                if total > 0:
                    result[name] = total

    logger.debug("技能表解析完成: %d 项技能", len(result))
    return result


# ------- 背景 -------

def _parse_background(ws) -> dict[str, str]:
    """提取七大背景项（AA 列，可跳过）"""
    bg: dict[str, str] = {}
    for field, row in _BG_ROWS.items():
        raw = ws.cell(row=row, column=_BG_COL).value
        val = str(raw).strip() if raw else ""
        if val:
            bg[field] = val
    return bg


# ------- 背包 -------

def _parse_inventory(ws) -> list[str]:
    """提取背包物品（F 列和 N 列的自由文本）"""
    items: list[str] = []
    for r in range(_ROW_INV_START, _ROW_INV_END + 1):
        for col in (_INV_COL_LEFT, _INV_COL_RIGHT):
            raw = ws.cell(row=r, column=col).value
            val = str(raw).strip() if raw else ""
            if val and val not in ("状态", "部位", "物品名称", "背包格↓", ""):
                items.append(val)
    return items


# ------- 法术 -------

def _parse_spells(ws) -> list[dict[str, str]]:
    """提取法术列表：名称 / 消耗 / 作用"""
    spells: list[dict[str, str]] = []
    for r in range(_SPELL_START, _SPELL_START + 20):
        name_raw = ws.cell(row=r, column=_SPELL_COL_NAME).value
        if not name_raw:
            continue
        name = str(name_raw).strip()
        if not name or name.startswith("编号"):
            continue
        cost = str(ws.cell(row=r, column=_SPELL_COST).value or "").strip()
        effect = str(ws.cell(row=r, column=_SPELL_EFFECT).value or "").strip()
        spells.append({"name": name, "cost": cost, "effect": effect})
    return spells


# ------- 恐惧症和躁狂症 -------

def _parse_phobias(ws) -> str:
    """提取恐惧症和躁狂症（AA75，部分角色卡可能为空）"""
    raw = ws.cell(row=_PHOBIA_ROW, column=_BG_COL).value
    return str(raw).strip() if raw else ""


# ------- 完整背景故事 -------

def _parse_full_backstory(ws) -> str:
    """提取 W77 的小说式完整背景故事（非七项摘要）"""
    raw = ws.cell(row=_FULL_BG_ROW, column=_FULL_BG_COL).value
    return str(raw).strip() if raw else ""
